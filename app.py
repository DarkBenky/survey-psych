import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import time
import os
import base64
from streamlit_js_eval import streamlit_js_eval


st.set_page_config(page_title='Test results calculator', page_icon='📝', layout='centered', initial_sidebar_state='collapsed')

    
class Answer:
    def __init__(self, category : str, value : int, question_number : int):
        self.category = category
        self.value = value
        self.question_number = question_number
        self.evaluated = False
        
    def __str__(self):
        return f'Category: {self.category}, Value: {self.value}, Question Number: {self.question_number}'

    def __repr__(self) -> str:
        return f'Category: {self.category}, Value: {self.value}, Question Number: {self.question_number}'
    
class Survey:
    def __init__(self, categories):
        self.categories = categories
        self.answers = {category: [] for category in categories}
        self.adhd_index = []
        self.wb = Workbook()
        self.ws = self.wb.active
        self.behavior = []
        self.opposition = []
        self.adhd_hyperactivity = {}
        self.adhd_inattention = {}
    
    def add_adhd_inattention(self, answer : Answer):
        self.adhd_inattention[int(answer.question_number)] = answer
        
    def calculate_adhd_inattention(self, age: int , type_of_test : str):
        if type_of_test == "Parent":
            score = 0
            if int(self.adhd_inattention[68].value) in [2,3] and int(self.adhd_inattention[79].value) in [2,3]:
                score += 1
                self.adhd_inattention[68].evaluated = True
                self.adhd_inattention[79].evaluated = True
            
            for answer in self.adhd_inattention.values():
                if int(answer.value) in [2,3] and int(answer.question_number) not in [68,79]:
                    answer.evaluated = True
                    score += 1
                elif int(answer.question_number) not in [68,79]:
                    answer.evaluated = False
                    
            if age <= 16:
                if score >= 6:
                    return score , True , [{
                        "Category": "ADHD Nepozornost",
                        "Value": answer.value,
                        "Question Number": answer.question_number,
                        "Evaluated": answer.evaluated
                        } for answer in self.adhd_inattention.values()]
                else:
                    return score , False , [{
                        "Category": "ADHD Nepozornost",
                        "Value": answer.value,
                        "Question Number": answer.question_number,
                        "Evaluated": answer.evaluated
                        } for answer in self.adhd_inattention.values()]
            else:
                if score >= 5:
                    return score , True , [{
                        "Category": "ADHD Nepozornost",
                        "Value": answer.value,
                        "Question Number": answer.question_number,
                        "Evaluated": answer.evaluated
                        } for answer in self.adhd_inattention.values()]
                else:
                    return score , False , [{
                        "Category": "ADHD Nepozornost",
                        "Value": answer.value,
                        "Question Number": answer.question_number,
                        "Evaluated": answer.evaluated
                        } for answer in self.adhd_inattention.values()]
                    
        elif type_of_test == "Self Evaluation":
            score = 0
            if int(self.adhd_inattention[31].value) in [2,3] or int(self.adhd_inattention[39].value) == 3:
                score += 1
                self.adhd_inattention[31].evaluated = True
                self.adhd_inattention[39].evaluated = True
                
            if int(self.adhd_inattention[61].value) in [2,3] and int(self.adhd_inattention[17].value) in [2,3]:
                score += 1
                self.adhd_inattention[61].evaluated = True
                self.adhd_inattention[17].evaluated = True
            
            
            for answer in self.adhd_inattention.values():
                if int(answer.value) == 3 and int(answer.question_number) in [21,51,77,32]:
                    answer.evaluated = True
                    score += 1
                elif int(answer.question_number) in [5,42,63] and int(answer.value) in [2,3]:
                    answer.evaluated = True
                    score += 1
                    
            if age <= 16:
                if score >= 6:
                    return score , True , [{
                        "Category": "ADHD Nepozornost",
                        "Value": answer.value,
                        "Question Number": answer.question_number,
                        "Evaluated": answer.evaluated
                        } for answer in self.adhd_inattention.values()]
                else:
                    return score , False , [{
                        "Category": "ADHD Nepozornost",
                        "Value": answer.value,
                        "Question Number": answer.question_number,
                        "Evaluated": answer.evaluated
                        } for answer in self.adhd_inattention.values()]
            else:
                if score >= 5:
                    return score , True , [{
                        "Category": "ADHD Nepozornost",
                        "Value": answer.value,
                        "Question Number": answer.question_number,
                        "Evaluated": answer.evaluated
                        } for answer in self.adhd_inattention.values()]
                else:
                    return score , False , [{
                        "Category": "ADHD Nepozornost",
                        "Value": answer.value,
                        "Question Number": answer.question_number,
                        "Evaluated": answer.evaluated
                        } for answer in self.adhd_inattention.values()]
            
    def add_adhd_hyperactivity(self, answer : Answer):
        self.adhd_hyperactivity[int(answer.question_number)] = answer
        
    def calculate_adhd_hyperactivity(self, age: int , type_of_test : str):
        if type_of_test == "Parent":
            score = 0
            if int(self.adhd_hyperactivity[69].value) in [2,3] and int(self.adhd_hyperactivity[99].value) in [2,3]:
                score += 1
                self.adhd_hyperactivity[69].evaluated = True
                self.adhd_hyperactivity[99].evaluated = True
                
            if int(self.adhd_hyperactivity[54].value) in [2,3] and int(self.adhd_hyperactivity[45].value) in [2,3]:
                score += 1
                self.adhd_hyperactivity[54].evaluated = True
                self.adhd_hyperactivity[45].evaluated = True
            
            for answer in self.adhd_hyperactivity.values():
                if int(answer.value) in [2,3] and int(answer.question_number) not in [69,99,54,45]:
                    answer.evaluated = True
                    score += 1
                elif int(answer.question_number) not in [69,99,54,45]:
                    answer.evaluated = False
                    
            if age <= 16:
                if score >= 6:
                    return score , True , [{
                        "Category": "ADHD Hyperactivity",
                        "Value": answer.value,
                        "Question Number": answer.question_number,
                        "Evaluated": answer.evaluated
                        } for answer in self.adhd_hyperactivity.values()]
                else:
                    return score , False , [{
                        "Category": "ADHD Hyperactivity",
                        "Value": answer.value,
                        "Question Number": answer.question_number,
                        "Evaluated": answer.evaluated
                        } for answer in self.adhd_hyperactivity.values()]
            else:
                if score >= 5:
                    return score , True , [{
                        "Category": "ADHD Hyperactivity",
                        "Value": answer.value,
                        "Question Number": answer.question_number,
                        "Evaluated": answer.evaluated
                        } for answer in self.adhd_hyperactivity.values()]
                else:
                    return score , False , [{
                        "Category": "ADHD Hyperactivity",
                        "Value": answer.value,
                        "Question Number": answer.question_number,
                        "Evaluated": answer.evaluated
                        } for answer in self.adhd_hyperactivity.values()]
                    
        elif type_of_test == "Self Evaluation":
            score = 0
            if int(self.adhd_hyperactivity[20].value) in [2,3] or int(self.adhd_hyperactivity[7].value) in [2,3]:
                score += 1
                self.adhd_hyperactivity[20].evaluated = True
                self.adhd_hyperactivity[7].evaluated = True
                
            if int(self.adhd_hyperactivity[66].value) in [2,3] or int(self.adhd_hyperactivity[55].value) == 3:
                score += 1
                self.adhd_hyperactivity[66].evaluated = True
                self.adhd_hyperactivity[55].evaluated = True
                
            for answer in self.adhd_hyperactivity.values():
                if int(answer.value) == 3 and int(answer.question_number) == 60:
                    answer.evaluated = True
                    score += 1
                elif int(answer.question_number) in [64,84,34,9,27,6] and int(answer.value) in [2,3]:
                    answer.evaluated = True
                    score += 1
                    
            if age <= 16:
                if score >= 6:
                    return score , True , [{
                        "Category": "ADHD Hyperactivity",
                        "Value": answer.value,
                        "Question Number": answer.question_number,
                        "Evaluated": answer.evaluated
                        } for answer in self.adhd_hyperactivity.values()]
                else:
                    return score , False , [{
                        "Category": "ADHD Hyperactivity",
                        "Value": answer.value,
                        "Question Number": answer.question_number,
                        "Evaluated": answer.evaluated
                        } for answer in self.adhd_hyperactivity.values()]
            else:
                if score >= 5:
                    return score , True , [{
                        "Category": "ADHD Hyperactivity",
                        "Value": answer.value,
                        "Question Number": answer.question_number,
                        "Evaluated": answer.evaluated
                        } for answer in self.adhd_hyperactivity.values()]
                else:
                    return score , False , [{
                        "Category": "ADHD Hyperactivity",
                        "Value": answer.value,
                        "Question Number": answer.question_number,
                        "Evaluated": answer.evaluated
                        } for answer in self.adhd_hyperactivity.values()]
        
    def add_porucha_chovania(self, answer : Answer):
        self.behavior.append(answer)
    
    def add_porucha_vzdoru(self, answer):
        self.opposition.append(answer)
        
    def calculate_opposition(self , type_of_test : str):
        if type_of_test == "Parent":
            score = 0
            for answer in self.opposition:
                if int(answer.value) in [2,3]:
                    answer.evaluated = True
                    score += 1
                else:
                    answer.evaluated = False
        
            if score >= 4:
                return score , True , [{
                    "Category": "Porucha Vzdoru",
                    "Value": answer.value,
                    "Question Number": answer.question_number,
                    "Evaluated": answer.evaluated
                    } for answer in self.opposition]
            else:
                return score , False , [{
                    "Category": "Porucha Vzdoru",
                    "Value": answer.value,
                    "Question Number": answer.question_number,
                    "Evaluated": answer.evaluated
                    } for answer in self.opposition]
                
        elif type_of_test == "Self Evaluation":
            score = 0
            for answer in self.opposition:
                if int(answer.question_number) in [67,1] and int(answer.value) == 3:
                    answer.evaluated = True
                    score += 1
                elif int(answer.value) in [2,3]:
                    answer.evaluated = False
        
            if score >= 4:
                return score , True , [{
                    "Category": "Porucha Vzdoru",
                    "Value": answer.value,
                    "Question Number": answer.question_number,
                    "Evaluated": answer.evaluated
                    } for answer in self.opposition]
            else:
                return score , False , [{
                    "Category": "Porucha Vzdoru",
                    "Value": answer.value,
                    "Question Number": answer.question_number,
                    "Evaluated": answer.evaluated
                    } for answer in self.opposition]
        
    def calculate_behavior(self , type_of_test : str):
        if type_of_test == "Parent":    
            score = 0
            for answer in self.behavior:
                if int(answer.question_number) in [16,30,56,91,6]:
                    if int(answer.value) in [2,3]:
                        answer.evaluated = True
                        score += 1
                    else:
                        answer.evaluated = False
                else:
                    if int(answer.value) in [1,2,3]:
                        answer.evaluated = True
                        score += 1
                    else:
                        answer.evaluated = False
                        
            if score >= 4:
                return score , True , [{ 
                    "Category": "Porucha Chovania",
                    "Value": answer.value,
                    "Question Number": answer.question_number,
                    "Evaluated": answer.evaluated
                    } for answer in self.behavior]
            else:
                return score , False , [{
                    "Category": "Porucha Chovania",
                    "Value": answer.value,
                    "Question Number": answer.question_number,
                    "Evaluated" : answer.evaluated
                    } for answer in self.behavior]
                
        elif type_of_test == "Self Evaluation":
            score = 0
            for answer in self.behavior:
                if int(answer.question_number) in [25,38,72,16,91,33] and int(answer.value) in [2,3]:
                    answer.evaluated = True
                    score += 1
                else:
                    if int(answer.value) in [1,2,3]:
                        answer.evaluated = True
                        score += 1
                        
            if score >= 3:
                return score , True , [{ 
                    "Category": "Porucha Chovania",
                    "Value": answer.value,
                    "Question Number": answer.question_number,
                    "Evaluated": answer.evaluated
                    } for answer in self.behavior]
            else:
                return score , False , [{
                    "Category": "Porucha Chovania",
                    "Value": answer.value,
                    "Question Number": answer.question_number,
                    "Evaluated" : answer.evaluated
                    } for answer in self.behavior]
                
          
    def add_adhd_index(self, answer):
        self.adhd_index.append(answer)
        
    def represent_adhd_index(self):
        return [
            {
            "Category": "ADHD",
            "Value": answer.value,
            "Question Number": answer.question_number,
            "Evaluated": "true" if answer.evaluated else "false"
            }
            for answer in self.adhd_index]
        
    def calculate_adhd_index(self , type_of_test : str):
        if type_of_test == "Parent":
            overall_score = 0
            for answer in self.adhd_index:
                if int(answer.question_number) in [19,47,67,84,99,101,104]:
                    if int(answer.value) in [0,1]:
                        answer.evaluated = False
                    else:
                        answer.evaluated = True
                        overall_score += int(answer.value) - 1
                else:
                    if int(answer.value) in [0,1]:
                        answer.evaluated = False
                    else:
                        answer.evaluated = True
                        overall_score += 2
            
            if overall_score == 0:
                self.probability = 11
            elif overall_score == 1:
                self.probability = 29
            elif overall_score == 2:
                self.probability = 41
            elif overall_score == 3:
                self.probability = 51
            elif overall_score == 4:
                self.probability = 56
            elif overall_score == 5:
                self.probability = 64
            elif overall_score == 6:
                self.probability = 71
            elif overall_score == 7:
                self.probability = 77
            elif overall_score == 8:
                self.probability = 82
            elif overall_score == 9:
                self.probability = 87
            elif overall_score == 10:
                self.probability = 91
            elif overall_score == 11:
                self.probability = 94
            elif overall_score == 12:
                self.probability = 97
            elif overall_score == 13:
                self.probability = 98
            elif overall_score >= 14:
                self.probability = 99
                
        elif type_of_test == "Self Evaluation":
            overall_score = 0
            for answer in self.adhd_index:
                if int(answer.question_number) in [6,9,15,21,43,63]:
                    if int(answer.value) in [0,1]:
                        answer.evaluated = False
                    else:
                        answer.evaluated = True
                        overall_score += int(answer.value) - 1
                else:
                    if int(answer.value) in [0,1]:
                        answer.evaluated = False
                    else:
                        answer.evaluated = True
                        if int(answer.question_number) in [34,50]:
                            overall_score += 1
                        else:
                            overall_score += 2
            
            if overall_score == 0:
                self.probability = 26
            elif overall_score == 1:
                self.probability = 35
            elif overall_score == 2:
                self.probability = 44
            elif overall_score == 3:
                self.probability = 52
            elif overall_score == 4:
                self.probability = 59
            elif overall_score == 5:
                self.probability = 66
            elif overall_score == 6:
                self.probability = 73
            elif overall_score == 7:
                self.probability = 78
            elif overall_score == 8:
                self.probability = 83
            elif overall_score == 9:
                self.probability = 87
            elif overall_score == 10:
                self.probability = 91
            elif overall_score == 11:
                self.probability = 94
            elif overall_score == 12:
                self.probability = 96
            elif overall_score == 13:
                self.probability = 97
            elif overall_score in  [14, 15]:
                self.probability = 98
            elif overall_score >= 16:
                self.probability = 99
        
        return overall_score , self.probability
            
        
    def add_answer(self, answer):
        self.answers[answer.category].append(answer)
    
    def calculate_score(self):
        scores = {category: 0 for category in self.categories}
        for category, answers in self.answers.items():
            for answer in answers:
                scores[category] += int(answer.value)
        return scores
    
    def output(self):
        answers_str = {category: [answer for answer in answers] for category, answers in self.answers.items()}
        scores = self.calculate_score()
        for category, score in scores.items():
            answers_str[category].append({'Score': score})
        return answers_str
    
    def rgb_to_hex(self, rgb):
        r, g, b = rgb.split(',')
        return f'{int(r):02x}{int(g):02x}{int(b):02x}'
        
    def export_to_excel(self, name , age , sex , description , type_of_test: str):
        self.ws.append(["Type of Test" , type_of_test ,"Name :" , name , "Age :" , age , "Sex :" , sex , "Description :" , description , "Time of creation :" , time.ctime()])
        # Set fill color for the entire row
        for cell in self.ws[self.ws.max_row]:
            cell.fill = PatternFill(start_color=self.rgb_to_hex("130,235,144"), end_color=self.rgb_to_hex("235,210,130"), fill_type="solid")
        
        print("test:",type_of_test)
        score_adhd_index , probability_adhd = self.calculate_adhd_index(type_of_test)
        self.ws.append(["ADHD Index score :" , score_adhd_index , "Probability of ADHD :" , str(probability_adhd)+' %'])
        for cell in self.ws[self.ws.max_row]:
            cell.fill = PatternFill(start_color=self.rgb_to_hex("235,210,130"), end_color=self.rgb_to_hex("235,210,130"), fill_type="solid")
        for answer in self.adhd_index:
            self.ws.append(["Question number :", answer.question_number, "Answer value :", answer.value])
            # add color to answer value cell based on the value
            if answer.value == '0':
                self.ws.cell(row=self.ws.max_row, column=4).fill = PatternFill(start_color=self.rgb_to_hex("58,67,180"), end_color=self.rgb_to_hex("58,67,180"), fill_type = "solid")
            elif answer.value == '1':
                self.ws.cell(row=self.ws.max_row, column=4).fill = PatternFill(start_color=self.rgb_to_hex("253,29,233"), end_color=self.rgb_to_hex("253,29,233"), fill_type = "solid")
            elif answer.value == '2':
                self.ws.cell(row=self.ws.max_row, column=4).fill = PatternFill(start_color=self.rgb_to_hex("253,67,67"), end_color=self.rgb_to_hex("253,67,67"), fill_type = "solid")
            elif answer.value == '3':
                self.ws.cell(row=self.ws.max_row, column=4).fill = PatternFill(start_color=self.rgb_to_hex("252,207,69"), end_color=self.rgb_to_hex("252,207,69"), fill_type = "solid")
         
        self.ws.append(["ADHD Hyperactivity score :" , self.calculate_adhd_hyperactivity(age=age , type_of_test=type_of_test)[0] , "ADHD Hyperactivity :" , self.calculate_adhd_hyperactivity(age=age, type_of_test=type_of_test)[1]])
        for cell in self.ws[self.ws.max_row]:
            cell.fill = PatternFill(start_color=self.rgb_to_hex("235,210,130"), end_color=self.rgb_to_hex("235,210,130"), fill_type="solid")
            
        for answer in self.adhd_hyperactivity.values():
            evaluated = 'true' if answer.evaluated else 'false'
            self.ws.append(["Question number :", answer.question_number, "Answer value :", answer.value , "Evaluated :" , evaluated])
            # add color to answer value cell based on the value
            if answer.evaluated == True:
                self.ws.cell(row=self.ws.max_row, column=4).fill = PatternFill(start_color=self.rgb_to_hex("58,67,180"), end_color=self.rgb_to_hex("58,67,180"), fill_type = "solid")
            else:
                self.ws.cell(row=self.ws.max_row, column=4).fill = PatternFill(start_color=self.rgb_to_hex("253,29,233"), end_color=self.rgb_to_hex("253,29,233"), fill_type = "solid")
            if type_of_test == "Parent":
                if answer.question_number in [69,99]:
                    self.ws.cell(row=self.ws.max_row, column=1).fill = PatternFill(start_color=self.rgb_to_hex("252,207,69"), end_color=self.rgb_to_hex("252,207,69"), fill_type = "solid")    
                if answer.question_number in [54,45]:
                    self.ws.cell(row=self.ws.max_row, column=1).fill = PatternFill(start_color=self.rgb_to_hex("253,67,67"), end_color=self.rgb_to_hex("253,67,67"), fill_type = "solid")
            elif type_of_test == "Self Evaluation":
                if answer.question_number in [20,7]:
                    self.ws.cell(row=self.ws.max_row, column=1).fill = PatternFill(start_color=self.rgb_to_hex("252,207,69"), end_color=self.rgb_to_hex("252,207,69"), fill_type = "solid")
                if answer.question_number in [66,55]:
                    self.ws.cell(row=self.ws.max_row, column=1).fill = PatternFill(start_color=self.rgb_to_hex("253,67,67"), end_color=self.rgb_to_hex("253,67,67"), fill_type = "solid")
                
        self.ws.append(["ADHD Inattention score :" , self.calculate_adhd_inattention(age=age , type_of_test=type_of_test)[0] , "ADHD Inattention :" , self.calculate_adhd_inattention(age=age , type_of_test=type_of_test)[1]])
        for cell in self.ws[self.ws.max_row]:
            cell.fill = PatternFill(start_color=self.rgb_to_hex("235,210,130"), end_color=self.rgb_to_hex("235,210,130"), fill_type="solid")
        for answer in self.adhd_inattention.values():
            evaluated = 'true' if answer.evaluated else 'false'
            self.ws.append(["Question number :", answer.question_number, "Answer value :", answer.value , "Evaluated :" , evaluated])
            # add color to answer value cell based on the value
            if answer.evaluated == True:
                self.ws.cell(row=self.ws.max_row, column=4).fill = PatternFill(start_color=self.rgb_to_hex("58,67,180"), end_color=self.rgb_to_hex("58,67,180"), fill_type = "solid")
            else:
                self.ws.cell(row=self.ws.max_row, column=4).fill = PatternFill(start_color=self.rgb_to_hex("253,29,233"), end_color=self.rgb_to_hex("253,29,233"), fill_type = "solid")
            if type_of_test == "Parent":
                if answer.question_number in [68,79]:
                    self.ws.cell(row=self.ws.max_row, column=1).fill = PatternFill(start_color=self.rgb_to_hex("252,207,69"), end_color=self.rgb_to_hex("252,207,69"), fill_type = "solid")
            elif type_of_test == "Self Evaluation":
                if answer.question_number in [31,39]:
                    self.ws.cell(row=self.ws.max_row, column=1).fill = PatternFill(start_color=self.rgb_to_hex("252,207,69"), end_color=self.rgb_to_hex("252,207,69"), fill_type = "solid")
                if answer.question_number in [61,17]:
                    self.ws.cell(row=self.ws.max_row, column=1).fill = PatternFill(start_color=self.rgb_to_hex("253,67,67"), end_color=self.rgb_to_hex("253,67,67"), fill_type = "solid")
            
        self.ws.append(["Oppositional defiant disorder score :" , self.calculate_opposition(type_of_test=type_of_test)[0] , "Oppositional defiant disorder :" , self.calculate_opposition(type_of_test=type_of_test)[1]])
        for cell in self.ws[self.ws.max_row]:
            cell.fill = PatternFill(start_color=self.rgb_to_hex("235,210,130"), end_color=self.rgb_to_hex("235,210,130"), fill_type="solid")
        for answer in self.opposition:
            evaluated = 'true' if answer.evaluated else 'false'
            self.ws.append(["Question number :", answer.question_number, "Answer value :", answer.value , "Evaluated :" , evaluated])
            # add color to answer value cell based on the value
            if answer.evaluated == True:
                self.ws.cell(row=self.ws.max_row, column=4).fill = PatternFill(start_color=self.rgb_to_hex("58,67,180"), end_color=self.rgb_to_hex("58,67,180"), fill_type = "solid")
            else:
                self.ws.cell(row=self.ws.max_row, column=4).fill = PatternFill(start_color=self.rgb_to_hex("253,29,233"), end_color=self.rgb_to_hex("253,29,233"), fill_type = "solid")

        self.ws.append(["Behavior disorder score :" , self.calculate_behavior(type_of_test=type_of_test)[0] , "Behavior disorder :" , self.calculate_behavior(type_of_test=type_of_test)[1]])
        for cell in self.ws[self.ws.max_row]:
            cell.fill = PatternFill(start_color=self.rgb_to_hex("235,210,130"), end_color=self.rgb_to_hex("235,210,130"), fill_type="solid")
        for answer in self.behavior:
            evaluated = 'true' if answer.evaluated else 'false'
            self.ws.append(["Question number :", answer.question_number, "Answer value :", answer.value , "Evaluated :" , evaluated])
            # add color to answer value cell based on the value
            if answer.evaluated == True:
                self.ws.cell(row=self.ws.max_row, column=4).fill = PatternFill(start_color=self.rgb_to_hex("58,67,180"), end_color=self.rgb_to_hex("58,67,180"), fill_type = "solid")
            else:
                self.ws.cell(row=self.ws.max_row, column=4).fill = PatternFill(start_color=self.rgb_to_hex("253,29,233"), end_color=self.rgb_to_hex("253,29,233"), fill_type = "solid")
                       
        
        for category, answers in self.answers.items():
            self.ws.append(["Category :" , category , "Overall score :" , self.calculate_score()[category]])
            # Set fill color for the entire row
            for cell in self.ws[self.ws.max_row]:
                cell.fill = PatternFill(start_color=self.rgb_to_hex("235,210,130"), end_color=self.rgb_to_hex("235,210,130"), fill_type="solid")
            for answer in answers:
                self.ws.append(["Question number :", answer.question_number, "Answer value :", answer.value, ])
                # add color to answer value cell based on the value
                if answer.value == '0':
                    self.ws.cell(row=self.ws.max_row, column=4).fill = PatternFill(start_color=self.rgb_to_hex("58,67,180"), end_color=self.rgb_to_hex("58,67,180"), fill_type = "solid")
                elif answer.value == '1':
                    self.ws.cell(row=self.ws.max_row, column=4).fill = PatternFill(start_color=self.rgb_to_hex("253,29,233"), end_color=self.rgb_to_hex("253,29,233"), fill_type = "solid")
                elif answer.value == '2':
                    self.ws.cell(row=self.ws.max_row, column=4).fill = PatternFill(start_color=self.rgb_to_hex("253,67,67"), end_color=self.rgb_to_hex("253,67,67"), fill_type = "solid")
                elif answer.value == '3':
                    self.ws.cell(row=self.ws.max_row, column=4).fill = PatternFill(start_color=self.rgb_to_hex("252,207,69"), end_color=self.rgb_to_hex("252,207,69"), fill_type = "solid")
        
        if os.path.exists('temp.xlsx'):
            os.remove('temp.xlsx')
        self.wb.save('temp.xlsx')
        self.wb = Workbook()
        self.ws = self.wb.active
        
    
    
def main_page():
    answers = {}

    st.title('Test results calculator')
    
    st.write('Enter your details')
    # Create columns for the grid layout
    c1, c2, c3, c4 , c5 = st.columns(5)
    # Create a container for each column
    with c1:
        name = st.text_input('Name')
    with c2:
        age = st.number_input('Age', step=1, value=18)
    with c3:
        sex = st.radio("Sex", options=["Female","Male"])
    with c4:
        description = st.text_area('Description')
    with c5:
        type_of_test = st.radio("Type of test", options=["Parent","Self Evaluation"])

    form = st.form('form')
        
    with form:
        try:
            if type_of_test == "Parent": 
                if screen_width > 720: 
                    # Create columns for the grid layout
                    col1, col2, col3, col4 = st.columns(4)
                    # Create a container for each column
                    with col1:
                        for i in range(0, 27):
                            question_number = i * 4 + 1
                            st.markdown('---')
                            answer = st.radio(f'Select an answer for question {question_number}:', ['0', '1', '2', '3'], key=f'radio_{question_number}')
                            answers[question_number] = answer

                    with col2:
                        for i in range(0, 27):
                            question_number = i * 4 + 2
                            st.markdown('---')
                            answer = st.radio(f'Select an answer for question {question_number}:', ['0', '1', '2', '3'], key=f'radio_{question_number}')
                            answers[question_number] = answer

                    with col3:
                        for i in range(0, 27):
                            question_number = i * 4 + 3
                            st.markdown('---')
                            answer = st.radio(f'Select an answer for question {question_number}:', ['0', '1', '2', '3'], key=f'radio_{question_number}')
                            answers[question_number] = answer

                    with col4:
                        for i in range(0, 27):
                            question_number = i * 4 + 4
                            st.markdown('---')
                            answer = st.radio(f'Select an answer for question {question_number}:', ['0', '1', '2', '3'], key=f'radio_{question_number}')
                            answers[question_number] = answer

                    

                else:
                    for i in range(0, 108):
                        st.markdown('---')
                        answer = st.radio(f'Select an answer for question {i + 1}:', ['0', '1', '2', '3'], key=f'radio_{i}')
                        answers[i + 1] = answer
                st.markdown('---')
            else:
                if screen_width > 720:
                    # Create columns for the grid layout
                    col1, col2, col3, col4 = st.columns(4)
                    # Create a container for each column
                    with col1:
                        for i in range(0, 25):
                            question_number = i * 4 + 1
                            st.markdown('---')
                            answer = st.radio(f'Select an answer for question {question_number}:', ['0', '1', '2', '3'], key=f'radio_{question_number}')
                            answers[question_number] = answer

                    with col2:
                        for i in range(0, 24):
                            question_number = i * 4 + 2
                            st.markdown('---')
                            answer = st.radio(f'Select an answer for question {question_number}:', ['0', '1', '2', '3'], key=f'radio_{question_number}')
                            answers[question_number] = answer

                    with col3:
                        for i in range(0, 24):
                            question_number = i * 4 + 3
                            st.markdown('---')
                            answer = st.radio(f'Select an answer for question {question_number}:', ['0', '1', '2', '3'], key=f'radio_{question_number}')
                            answers[question_number] = answer

                    with col4:
                        for i in range(0, 24):
                            question_number = i * 4 + 4
                            st.markdown('---')
                            answer = st.radio(f'Select an answer for question {question_number}:', ['0', '1', '2', '3'], key=f'radio_{question_number}')
                            answers[question_number] = answer
                else:
                    for i in range(0, 97):
                        st.markdown('---')
                        answer = st.radio(f'Select an answer for question {i + 1}:', ['0', '1', '2', '3'], key=f'radio_{i}')
                        answers[i + 1] = answer
                st.markdown('---')
        except:
            pass
        

        if st.form_submit_button('Submit'):
            survey = create_survey(answers , type_of_test)
            survey.export_to_excel(name , age , sex , description , type_of_test)
            calculate_score(survey , age , type_of_test)    
            

@st.cache_data()
def convert_to_bites(file_path):
    with open(file_path, 'rb') as file:
        return file.read()    
     
@st.cache_resource()    
def create_survey(answers , type_of_test):
    categories_page_one = ['IN' , 'HY' , 'LP' , 'EF' , 'AG' , 'PR' , 'GI' , 'AN' , 'AH' , 'CD' , 'OD' , 'PI' , 'NI',]
    survey = Survey(categories_page_one)

    for question_number , answer in answers.items():
        question_number = int(question_number)
        
        # Parent test
        if question_number in [47,95,35,68,79,84,28,97,101,2] and type_of_test == "Parent":
            survey.add_adhd_inattention(Answer('ADHD nepozornost' , answer , question_number))
        if question_number in [98,93,69,99,71,54,45,3,43,61,104] and type_of_test == "Parent":
            survey.add_adhd_hyperactivity(Answer('ADHD hyperactivity' , answer , question_number))
        if question_number in [16,30,27,39,41,96,11,78,65,89,56,58,91,76,6] and type_of_test == "Parent":
            survey.add_porucha_chovania(Answer('Porucha Chovania' , answer , question_number))
        if question_number in [14,73,48,102,94,59,21,57] and type_of_test == "Parent":
            survey.add_porucha_vzdoru(Answer('Porucha Vzdoru' , answer , question_number))
        if question_number in [19,35,47,67,84,88,98,99,101,104] and type_of_test == "Parent":
            survey.add_adhd_index(Answer('ADHD' , answer , question_number))

        # Self evaluation test
        if question_number in [31,39,63,42,61,17,21,51,5,77,32] and type_of_test == "Self Evaluation":
            survey.add_adhd_inattention(Answer('ADHD nepozornost' , answer , question_number))
        if question_number in [60,64,20,7,84,66,55,34,9,27,6] and type_of_test == "Self Evaluation":
            survey.add_adhd_hyperactivity(Answer('ADHD hyperactivity' , answer , question_number))
        if question_number in [25,38,59,86,47,13,72,82,78,16,52,91,8,33] and type_of_test == "Self Evaluation":
            survey.add_porucha_chovania(Answer('Porucha Chovania' , answer , question_number))
        if question_number in [67,74,87,24,1,3,62,94] and type_of_test == "Self Evaluation":
            survey.add_porucha_vzdoru(Answer('Porucha Vzdoru' , answer , question_number))
        if question_number in [6,9,15,21,34,35,43,50,61,63] and type_of_test == "Self Evaluation":
            survey.add_adhd_index(Answer('ADHD' , answer , question_number))
        
        if question_number in [1,8,18,26,32,42]:
            survey.add_answer(Answer('NI' , answer , question_number))
        if question_number in [31,33,38]:
            survey.add_answer(Answer('PI' , answer , question_number))
        if question_number in [14,21,48,57,59]:
            survey.add_answer(Answer('OD' , answer , question_number))
        if question_number in [6,11,16,27,30,39,41,56,58]:
            survey.add_answer(Answer('CD' , answer , question_number))
        if question_number in [3,43,45,54,61]:
            survey.add_answer(Answer('AH' , answer , question_number))
        if question_number in [2,28,35,47]:
            survey.add_answer(Answer('AN' , answer , question_number))
        if question_number in [19,25,29,34,40,50]:
            survey.add_answer(Answer('GI' , answer , question_number))
        if question_number in [10,13,24,62]:
            survey.add_answer(Answer('PR' , answer , question_number))
        if question_number in [16,22,27,30,39,46,48,57,58]:
            survey.add_answer(Answer('AG' , answer , question_number))
        if question_number in [34,37,63]:
            survey.add_answer(Answer('EF' , answer , question_number))
        if question_number in [5,7,9,15,36,51,53,60]:
            survey.add_answer(Answer('LP' , answer , question_number))
        if question_number in [19,43,45,50,52,54,55,61]:
            survey.add_answer(Answer('HY' , answer , question_number))
        if question_number in [12,23,28,44,47,49]:
            survey.add_answer(Answer('IN' , answer , question_number))
        if question_number in [74,80,105]:
            survey.add_answer(Answer('PI' , answer , question_number))
        if question_number in [73,94,102]:
            survey.add_answer(Answer('OD' , answer , question_number))
        if question_number in [65,76,78,89,91,96]:
            survey.add_answer(Answer('CD' , answer , question_number))
        if question_number in [69,71,93,98,99,104]:
            survey.add_answer(Answer('AH' , answer , question_number))
        if question_number in [68,79,84,95,97,101]:
            survey.add_answer(Answer('AN' , answer , question_number))
        if question_number in [67,81,85,99]:
            survey.add_answer(Answer('GI' , answer , question_number))
        if question_number in [64,92]:
            survey.add_answer(Answer('PR' , answer , question_number))
        if question_number in [65,83,86,94,102]:
            survey.add_answer(Answer('AG' , answer , question_number))
        if question_number in [72,75,79,84,90,97]:
            survey.add_answer(Answer('EF' , answer , question_number))
        if question_number in [87]:
            survey.add_answer(Answer('LP' , answer , question_number))
        if question_number in [69,71,93,98,99,104]:
            survey.add_answer(Answer('HY' , answer , question_number))
        if question_number in [67,77,88,95]:
            survey.add_answer(Answer('IN' , answer , question_number))
    
    return survey
    
def get_binary_file_downloader_link(file_path, file_label='File'):
    """
    Generates a link to download the given file.
    
    Parameters:
        file_path (str): The path to the file to be downloaded.
        file_label (str): The label for the download link.
        
    Returns:
        str: The HTML for the download link.
    """
    with open(file_path, 'rb') as f:
        bytes_data = f.read()
    b64 = base64.b64encode(bytes_data).decode()
    href = f'<a href="data:file/xlsx;base64,{b64}" download="{file_path}">{file_label}</a>'
    return href
    
# @st.cache_data()
def calculate_score(survey : Survey , age : int , type_of_test : str):        
    output = survey.output()
    adhd_score , adhd_probability = survey.calculate_adhd_index(type_of_test=type_of_test)
    behavior_disorder_value , behavior_disorder , answer_behavior = survey.calculate_behavior(type_of_test=type_of_test)
    oppositional_defiant_disorder_value , oppositional_defiant_disorder , answer_opposition = survey.calculate_opposition(type_of_test=type_of_test)
    adhd_answers = survey.represent_adhd_index()
    hyperactivity_score , hyperactivity_value , answer_hyperactivity = survey.calculate_adhd_hyperactivity(age=age, type_of_test=type_of_test)
    inattention_score , inattention_value , answer_inattention = survey.calculate_adhd_inattention(age=age , type_of_test=type_of_test)
    
    # download temp.xlsx file
    st.markdown(get_binary_file_downloader_link('temp.xlsx', 'Download Excel File'), unsafe_allow_html=True)
    
    st.header('ADHD Index')
    st.write("ADHD Index Score: ", adhd_score)
    st.write(f"Probability of ADHD: " , f"{adhd_probability} %")
    st.progress(adhd_probability,text=f"{len(adhd_answers)} / {adhd_score}")
    expand_adhd = st.expander(f'Answers for category ADHD Index', expanded=False)
    with expand_adhd:
        for answer in adhd_answers:
            st.markdown(f'Question Number: {answer["Question Number"]} , Value: {answer["Value"]}')
    st.markdown('---')
    
    st.header("ADHD Hyperactivity")
    st.write("Score ADHD Hyperactivity: ", hyperactivity_score)
    st.write("ADHD Hyperactivity: ", hyperactivity_value)
    st.progress(hyperactivity_score / (len(answer_hyperactivity) - 2), text=f"{hyperactivity_score} / {(len(answer_hyperactivity) - 2)}")
    expand_hyperactivity = st.expander(f'Answers for category ADHD Hyperactivity', expanded=False)
    with expand_hyperactivity:
        for answer in answer_hyperactivity:
            st.markdown(f'Question Number: {answer["Question Number"]} , Value: {answer["Value"]} , Evaluated: {answer["Evaluated"]}')
    st.markdown('---')
    
    st.header("ADHD Inattention")
    st.write("Score ADHD Inattention: ", inattention_score)
    st.write("ADHD Inattention: ", inattention_value)
    if type_of_test == "Parent":
        st.progress(inattention_score / (len(answer_inattention) -1) , text=f"{inattention_score} / {(len(answer_inattention) - 1)}")
    else:
        st.progress(inattention_score / (len(answer_inattention) - 2) , text=f"{inattention_score} / {(len(answer_inattention) - 2)}")
    expand_inattention = st.expander(f'Answers for category ADHD Inattention', expanded=False)
    with expand_inattention:
        for answer in answer_inattention:
            st.markdown(f'Question Number: {answer["Question Number"]} , Value: {answer["Value"]} , Evaluated: {answer["Evaluated"]}')
    st.markdown('---')
    
    st.header("Oppositional Defiant Disorder")
    st.write("Score Oppositional Defiant Disorder: ",  oppositional_defiant_disorder_value)
    st.write("Oppositional Defiant Disorder: ",  oppositional_defiant_disorder)
    st.progress(oppositional_defiant_disorder_value / len(answer_opposition), text=f"{oppositional_defiant_disorder_value} / {len(answer_opposition)}")
    expand_opposition = st.expander(f'Answers for category Oppositional Defiant Disorder', expanded=False)
    with expand_opposition:
        for answer in answer_opposition:
            st.markdown(f'Question Number: {answer["Question Number"]} , Value: {answer["Value"]} , Evaluated: {answer["Evaluated"]}')
    st.markdown('---')
    
    st.header("Behavior Disorder")
    st.write("Score for Behavior Disorder: ", behavior_disorder_value)
    st.write("Behavior Disorder: ", behavior_disorder)
    expand_behavior = st.expander(f'Answers for category Behavior Disorder', expanded=False)
    st.progress(behavior_disorder_value / len(answer_behavior), text=f"{behavior_disorder_value} / {len(answer_behavior)}")
    with expand_behavior:
        for answer in answer_behavior:
            st.markdown(f'Question Number: {answer["Question Number"]} , Value: {answer["Value"]} , Evaluated: {answer["Evaluated"]}')
    st.markdown('---')
    
        
    for category , element in output.items():
        st.header(f'Category: {category}')
        st.write("Score for Category: ", element[-1]['Score'])
        try :
            normalized_score = element[-1]['Score'] / ((len(element) - 1) * 3)
        except ZeroDivisionError:
            normalized_score = 0
        st.progress(normalized_score , text=f"{element[-1]['Score']} / {len(element) * 3}")
        expand = st.expander(f'Answers for category {category}', expanded=False)
        with expand:
            for answer in element[:-1]:
                st.markdown(f'Question Number: {answer.question_number} , Value: {answer.value}')
         
def get_screen_width():
    return streamlit_js_eval(js_expressions='screen.width', key = 'SCR')
    
if __name__ == '__main__':
    global screen_width
    screen_width = get_screen_width()
    main_page()
